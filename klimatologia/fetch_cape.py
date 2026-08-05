import urllib.request
import json
import time
import os
import openpyxl
from datetime import datetime

def fetch_thermodynamics(lat, lon, date_str):
    url = (
        f"https://historical-forecast-api.open-meteo.com/v1/forecast"
        f"?latitude={lat}&longitude={lon}"
        f"&start_date={date_str}&end_date={date_str}"
        f"&hourly=cape,lifted_index"
        f"&timezone=Europe/Warsaw"
    )
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=15) as response:
            data = json.loads(response.read().decode('utf-8'))
            
            cape_vals = [x for x in data['hourly']['cape'] if x is not None]
            li_vals = [x for x in data['hourly']['lifted_index'] if x is not None]
            
            max_cape = round(max(cape_vals)) if cape_vals else 0
            min_li = round(min(li_vals), 1) if li_vals else 0
            
            return max_cape, min_li
    except Exception as e:
        print(f"Błąd dla {date_str}: {e}")
        return None, None

def get_coords(location_str):
    # Słownik lokalizacji (skopiowany z poprzedniego skryptu)
    coords_dict = {
        "Zagwiździe": (50.85, 17.93), "Opole": (50.67, 17.92), "Prószków": (50.58, 17.87),
        "Murów": (50.86, 17.94), "Wrocław": (51.11, 17.04), "Paryż (FR)": (48.85, 2.35),
        "Jaśkowice": (50.60, 17.82), "Wilkowice": (50.77, 15.98), "Daniec": (50.63, 18.12),
        "Jurgów": (49.33, 20.14), "Warszawa": (52.23, 21.01), "Złotniki": (51.17, 16.97),
        "Gliwice": (50.28, 18.68), "Zator": (49.99, 19.43), "Biała": (50.38, 17.66),
        "Jastrzębie-Zdrój": (49.95, 18.58), "Kutno": (52.23, 19.36), "Piotrków Trybunalski": (51.41, 19.68),
        "Łosiów": (50.81, 17.58), "Lubliniec": (50.67, 18.68), "Poczesna": (50.75, 19.16),
        "Poręba": (50.48, 19.42), "Zawiercie": (50.48, 19.42), "Myszków": (50.58, 19.33),
        "Źlinice": (50.60, 17.94), "Prusice": (51.32, 16.83), "Ząbkowice Śląskie": (50.58, 16.82),
        "Ujazd": (50.38, 18.35), "Krapkowice": (50.48, 17.96), "Grodków": (50.70, 17.38),
        "Godzieszowa": (51.03, 16.83), "Świdnica": (50.85, 16.48), "Brzeg Dolny": (51.27, 16.72),
        "Ostrzeszów": (51.42, 17.92), "Lututów": (51.37, 18.43), "Siewierz": (50.48, 19.23),
        "[16 burz]": (49.33, 20.14)
    }
    for key, val in coords_dict.items():
        if key.lower() in location_str.lower():
            return val
    return (51.11, 17.04) # Default to Wrocław if unknown

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_file = os.path.join(script_dir, "Pamiętnik burz.xlsx")
    output_file = os.path.join(script_dir, "Pamiętnik_CAPE_LI.xlsx")
    
    print("Otwieranie Pamiętnika Burz (tylko do odczytu dat)...")
    wb_in = openpyxl.load_workbook(input_file, data_only=True)
    ws_in = wb_in["Burze"]
    
    # Tworzymy nowy mały skoroszyt tylko na pobrane dane, by nie zepsuć formatowań usera!
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Dane Termodynamiczne"
    
    ws_out.append(["L.p.", "Data", "Lokalizacja", "Max CAPE (J/kg)", "Min Lifted Index"])
    
    cache = {}
    
    for row in range(5, ws_in.max_row + 1):
        lp = ws_in.cell(row=row, column=3).value
        date_val = ws_in.cell(row=row, column=4).value
        loc = ws_in.cell(row=row, column=6).value
        
        if not date_val: continue
        
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val)[:10]
            
        coords = get_coords(str(loc))
        cache_key = f"{date_str}_{coords[0]}_{coords[1]}"
        
        if cache_key in cache:
            cape, li = cache[cache_key]
        else:
            print(f"Pobieranie CAPE dla {date_str} ({loc})...", end=" ")
            cape, li = fetch_thermodynamics(coords[0], coords[1], date_str)
            cache[cache_key] = (cape, li)
            print(f"-> CAPE: {cape}, LI: {li}")
            time.sleep(0.5) # Rate limit
            
        ws_out.append([lp, date_str, loc, cape, li])

    wb_out.save(output_file)
    print(f"Zapisano pobrane dane do {output_file}. Możesz je teraz skopiować do swojego pliku głównego!")

if __name__ == '__main__':
    main()
