import openpyxl
import os

def main():
    # Use the specific file mentioned by the user
    input_file = "Pamiętnik burz.xlsx"
    
    wb = openpyxl.load_workbook(input_file)
    ws = wb["Burze"]

    # L = 12 (Wiatr), J = 10 (Deszcz), K = 11 (Grad), M = 13 (Tornado)
    # N = 14 (Ocena zagrożeń), V = 22 (Ocena wyglądu), X = 24 (Opłacalność), W = 23 (Przejechane kilometry)
    # O to T = 15 to 20
    
    for r in range(5, ws.max_row + 1):
        f_wiatr = f"IF(L{r}>120,5,IF(L{r}>=100,4,IF(L{r}>=80,3,IF(L{r}>=60,2,IF(L{r}>=40,1,0)))))"
        f_opad = f"IF(J{r}>60,5,IF(J{r}>=45,4,IF(J{r}>=30,3,IF(J{r}>=15,2,IF(J{r}>=1,1,0)))))"
        f_grad = f"IF(K{r}>7,5,IF(K{r}>=5,4,IF(K{r}>=3,3,IF(K{r}>=1,2,IF(K{r}>0.1,1,0)))))"
        f_tor = f"IF(ISBLANK(M{r}),0,5)"
        
        f_z = f"({f_wiatr} + {f_opad} + {f_grad} + {f_tor})"
        
        # String representation for N
        formula_n = f'=IF({f_z}>15,"Ekstremalna (>15)",IF({f_z}>=12,"Bardzo silna (12-15)",IF({f_z}>=8,"Silna (8-11)",IF({f_z}>=4,"Umiarkowana (4-7)",IF({f_z}>=1,"Słaba (1-3)","Brak (0)")))))'
        ws.cell(row=r, column=14, value=formula_n)
        
        # Formula for Appearance Points (Wygląd)
        f_w = f"SUM(O{r}:T{r})"
        formula_v = f'=IF({f_w}>15,"Ekstremalna (>15)",IF({f_w}>=12,"Bardzo silna (12-15)",IF({f_w}>=8,"Silna (8-11)",IF({f_w}>=4,"Umiarkowana (4-7)",IF({f_w}>=1,"Słaba (1-3)","Brak (0)")))))'
        ws.cell(row=r, column=22, value=formula_v)
        
        # Formula for X (Opłacalność)
        formula_x = f'=IF(W{r}>0, ({f_z} + {f_w}) / W{r} * 100, "-")'
        ws.cell(row=r, column=24, value=formula_x)

    wb.save(input_file)
    print("Formulas applied successfully to", input_file)

if __name__ == "__main__":
    main()
