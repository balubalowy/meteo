import openpyxl
import os
import shutil

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_file = os.path.join(script_dir, "Pamiętnik burzdd.xlsx")
    output_file = os.path.join(script_dir, "Pamiętnik burz - poprawiony.xlsx")
    
    # Skopiuj odblokowany plik
    shutil.copyfile(input_file, output_file)
    
    wb = openpyxl.load_workbook(output_file)
    ws = wb["Burze"]

    for r in range(5, ws.max_row + 1):
        # Wiatr (L), Opad (J), Grad (K) - zabezpieczone funkcją ISNUMBER
        # Tornado (M) - odrzuci pustą komórkę lub znak "-"
        
        f_wiatr = f"IF(ISNUMBER(L{r}), IF(L{r}>120,5,IF(L{r}>=100,4,IF(L{r}>=80,3,IF(L{r}>=60,2,IF(L{r}>=40,1,0))))), 0)"
        f_opad = f"IF(ISNUMBER(J{r}), IF(J{r}>60,5,IF(J{r}>=45,4,IF(J{r}>=30,3,IF(J{r}>=15,2,IF(J{r}>=1,1,0))))), 0)"
        f_grad = f"IF(ISNUMBER(K{r}), IF(K{r}>7,5,IF(K{r}>=5,4,IF(K{r}>=3,3,IF(K{r}>=1,2,IF(K{r}>0.1,1,0))))), 0)"
        f_tor = f'IF(OR(ISBLANK(M{r}), M{r}="-", M{r}="brak", M{r}=0), 0, 5)'
        
        f_z = f"({f_wiatr} + {f_opad} + {f_grad} + {f_tor})"
        
        # N - Ocena zagrożeń
        formula_n = f'=IF({f_z}>15,"Ekstremalna (>15)",IF({f_z}>=12,"Bardzo silna (12-15)",IF({f_z}>=8,"Silna (8-11)",IF({f_z}>=4,"Umiarkowana (4-7)",IF({f_z}>=1,"Słaba (1-3)","Brak (0)")))))'
        ws.cell(row=r, column=14, value=formula_n)
        
        # V - Ocena wyglądu (SUM automatycznie ignoruje tekst, więc jest OK)
        f_w = f"SUM(O{r}:T{r})"
        formula_v = f'=IF({f_w}>15,"Ekstremalna (>15)",IF({f_w}>=12,"Bardzo silna (12-15)",IF({f_w}>=8,"Silna (8-11)",IF({f_w}>=4,"Umiarkowana (4-7)",IF({f_w}>=1,"Słaba (1-3)","Brak (0)")))))'
        ws.cell(row=r, column=22, value=formula_v)
        
        # X - Opłacalność
        formula_x = f'=IF(ISNUMBER(W{r}), IF(W{r}>0, ({f_z} + {f_w}) / W{r} * 100, "-"), "-")'
        ws.cell(row=r, column=24, value=formula_x)

    wb.save(output_file)
    print("Formulas fixed successfully in", output_file)

if __name__ == "__main__":
    main()
