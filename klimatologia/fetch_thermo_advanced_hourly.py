import urllib.request
import json
import time
import os
import openpyxl
import re
from datetime import datetime

def fetch_advanced_thermodynamics(lat, lon, date_str, storm_hour):
    url = (
        f"https://historical-forecast-api.open-meteo.com/v1/forecast"
        f"?latitude={lat}&longitude={lon}"
        f"&start_date={date_str}&end_date={date_str}"
        f"&hourly=cape,lifted_index,wind_speed_850hPa,wind_direction_850hPa,"
        f"wind_speed_700hPa,wind_direction_700hPa,wind_speed_500hPa,wind_direction_500hPa"
        f"&wind_speed_unit=kmh"
        f"&timezone=Europe/Warsaw"
    )
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=15) as response:
            data = json.loads(response.read().decode('utf-8'))
            hourly = data['hourly']
            
            # Bezpieczne parsowanie None na 0
            cape_vals = [x if x is not None else 0 for x in hourly['cape']]
            li_vals = [x if x is not None else 0 for x in hourly['lifted_index']]
            
            ws_850 = [x if x is not None else 0 for x in hourly['wind_speed_850hPa']]
            wd_850 = [x if x is not None else 0 for x in hourly['wind_direction_850hPa']]
            
            ws_700 = [x if x is not None else 0 for x in hourly['wind_speed_700hPa']]
            wd_700 = [x if x is not None else 0 for x in hourly['wind_direction_700hPa']]
            
            ws_500 = [x if x is not None else 0 for x in hourly['wind_speed_500hPa']]
            wd_500 = [x if x is not None else 0 for x in hourly['wind_direction_500hPa']]
            
            if not cape_vals:
                return None
                
            max_cape = max(cape_vals)
            min_li = min(li_vals)
            
            # Jeżeli storm_hour podana przez użytkownika istnieje i jest w zakresie 0-23
            if storm_hour is not None and 0 <= storm_hour <= 23:
                target_idx = storm_hour
                hour_source = "Wpisana godzina"
            else:
                # W przeciwnym razie fallback - godzina największego CAPE lub przepływu 500hPa
                if max_cape >= 100:
                    target_idx = cape_vals.index(max_cape)
                    hour_source = "Szczyt CAPE"
                else:
                    target_idx = ws_500.index(max(ws_500))
                    hour_source = "Szczyt Wiatru 500hPa"
                
            res = {
                "max_cape": round(max_cape),
                "min_li": round(min_li, 1),
                "ws_850": round(ws_850[target_idx]),
                "wd_850": round(wd_850[target_idx]),
                "ws_700": round(ws_700[target_idx]),
                "wd_700": round(wd_700[target_idx]),
                "ws_500": round(ws_500[target_idx]),
                "wd_500": round(wd_500[target_idx]),
                "target_hour": f"{target_idx:02d}:00",
                "source": hour_source
            }
            return res
            
    except Exception as e:
        print(f"Błąd dla {date_str}: {e}")
        return None

def get_coords(location_str):
    coords_dict = {
        "Zagwiździe": (50.85, 17.93), "Opole": (50.67, 17.92), "Prószków": (50.58, 17.87),
        "Murów": (50.86, 17.94), "Wrocław": (51.11, 17.04), "Paryż (FR)": (48.85, 2.35),
        "Jaśkowice": (50.60, 17.82), "Wilkowice": (50.77, 15.98), "Daniec": (50.63, 18.12),
        "Jurgów": (49.33, 20.14), "Warszawa": (52.23, 21.01), "Złotniki": (51.17, 16.97),
        "Gliwice": (50.28, 18.68), "Zator": (49.99, 19.43), "Biała": (50.38, 17.66),
        "Jastrzębie-Zdrój": (49.95, 18.58), "Kutno": (52.23, 19.36), "Piotrków Trybunalski": (51.41, 19.68),
        "Łosiów": (50.81, 17.58), "Lubliniec": (50.67, 18.68), "Poczesna": (50.75, 19.16),
        "Poręba": (50.48, 19.42), "Zawiercie": (50.48, 19.42), "Myszków": (50.58, 19.33),
        "Źlinice": (50.60, 17.94), "Prusice": (51.32, 16.83), "Ząbkowice Śląskie": (50.58, 16.82),
        "Ujazd": (50.38, 18.35), "Krapkowice": (50.48, 17.96), "Grodków": (50.70, 17.38),
        "Godzieszowa": (51.03, 16.83), "Świdnica": (50.85, 16.48), "Brzeg Dolny": (51.27, 16.72),
        "Ostrzeszów": (51.42, 17.92), "Lututów": (51.37, 18.43), "Siewierz": (50.48, 19.23),
        "[16 burz]": (49.33, 20.14)
    }
    for key, val in coords_dict.items():
        if key.lower() in location_str.lower():
            return val
    return (51.11, 17.04)

def parse_hour(time_str):
    if not time_str: return None
    time_str = str(time_str)
    # Wyciągamy pierwsze dwie cyfry (albo jedną) przed dwukropkiem, czyli godzinę startu
    match = re.search(r'(\d{1,2})', time_str)
    if match:
        return int(match.group(1))
    return None

def get_wind_dir_name(deg):
    dirs = ["N", "NE", "E", "SE", "S", "SW", "W", "NW", "N"]
    idx = round(deg / 45) % 8
    return dirs[idx]

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_file = os.path.join(script_dir, "Pamiętnik burz.xlsx")
    output_file = os.path.join(script_dir, "Pamiętnik_Kinematyka.xlsx")
    
    print("Otwieranie Pamiętnika Burz (odczyt dat i godzin)...")
    wb_in = openpyxl.load_workbook(input_file, data_only=True)
    ws_in = wb_in["Burze"]
    
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Dane Kinematyczne"
    
    headers = [
        "L.p.", "Data", "Godzina w dzienniku", "Lokalizacja", "Max CAPE (Doba)", "Min LI", 
        "Użyta Godzina dla Wiatru", "Źródło Godziny",
        "Kierunek Wiatru 850 hPa", "Kierunek Wiatru 700 hPa", 
        "Prędkość 850 hPa (km/h)", "Prędkość 700 hPa (km/h)", "Prędkość 500 hPa (km/h)"
    ]
    ws_out.append(headers)
    
    cache = {}
    
    for row in range(5, ws_in.max_row + 1):
        lp = ws_in.cell(row=row, column=3).value
        date_val = ws_in.cell(row=row, column=4).value
        time_val = ws_in.cell(row=row, column=5).value
        loc = ws_in.cell(row=row, column=6).value
        
        if not date_val: continue
        
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val)[:10]
            
        storm_hour = parse_hour(time_val)
        coords = get_coords(str(loc))
        
        # Klucz zależy teraz też od wybranej godziny
        cache_key = f"{date_str}_{coords[0]}_{coords[1]}_{storm_hour}"
        
        if cache_key in cache:
            res = cache[cache_key]
        else:
            print(f"Pobieranie kinematyki dla {date_str} ({loc}, G: {storm_hour})...", end=" ")
            res = fetch_advanced_thermodynamics(coords[0], coords[1], date_str, storm_hour)
            cache[cache_key] = res
            print("OK" if res else "BŁĄD")
            time.sleep(0.5)
            
        if res:
            dir_850_str = f"{res['wd_850']}° ({get_wind_dir_name(res['wd_850'])})"
            dir_700_str = f"{res['wd_700']}° ({get_wind_dir_name(res['wd_700'])})"
            
            ws_out.append([
                lp, date_str, str(time_val) if time_val else "Brak", loc, 
                res['max_cape'], res['min_li'], res['target_hour'], res['source'],
                dir_850_str, dir_700_str, 
                res['ws_850'], res['ws_700'], res['ws_500']
            ])

    wb_out.save(output_file)
    print(f"\nZapisano ulepszone dane do {output_file}.")

if __name__ == '__main__':
    main()
