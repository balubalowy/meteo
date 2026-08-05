import openpyxl
import os

def create_scale_sheet(wb):
    if "Skala" in wb.sheetnames:
        ws_skala = wb["Skala"]
        ws_skala.delete_rows(1, ws_skala.max_row)
    else:
        ws_skala = wb.create_sheet("Skala")
    
    # Nagłówki
    headers = [
        "Wiatr (km/h)", "Pkt", "", 
        "Opad (mm)", "Pkt", "", 
        "Grad (cm)", "Pkt", "", 
        "Kategoria ocen (Pkt)", "Nazwa oceny", "",
        "Tornado (obecność)"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        ws_skala.cell(row=1, column=col_idx, value=header)
        
    # Dane dla Wiatru (A, B)
    wiatr_data = [(0, 0), (40, 1), (60, 2), (80, 3), (100, 4), (120, 5)]
    for i, (val, pkt) in enumerate(wiatr_data, 2):
        ws_skala.cell(row=i, column=1, value=val)
        ws_skala.cell(row=i, column=2, value=pkt)
        
    # Dane dla Opadu (D, E)
    opad_data = [(0, 0), (1, 1), (15, 2), (30, 3), (45, 4), (60, 5)]
    for i, (val, pkt) in enumerate(opad_data, 2):
        ws_skala.cell(row=i, column=4, value=val)
        ws_skala.cell(row=i, column=5, value=pkt)

    # Dane dla Gradu (G, H)
    grad_data = [(0, 0), (0.1, 1), (1, 2), (3, 3), (5, 4), (7, 5)]
    for i, (val, pkt) in enumerate(grad_data, 2):
        ws_skala.cell(row=i, column=7, value=val)
        ws_skala.cell(row=i, column=8, value=pkt)
        
    # Skala kategorii ogólnej (J, K)
    kat_data = [
        (0, "Brak (0)"),
        (1, "Słaba (1-3)"),
        (4, "Umiarkowana (4-7)"),
        (8, "Silna (8-11)"),
        (12, "Bardzo silna (12-15)"),
        (16, "Ekstremalna (>15)")
    ]
    for i, (val, name) in enumerate(kat_data, 2):
        ws_skala.cell(row=i, column=10, value=val)
        ws_skala.cell(row=i, column=11, value=name)
        
    # Tornado punkty (M, N)
    ws_skala.cell(row=2, column=13, value="Punkty dodawane:")
    ws_skala.cell(row=2, column=14, value=5)
    
    # Ustawienie szerokości kolumn dla czytelności
    ws_skala.column_dimensions['A'].width = 15
    ws_skala.column_dimensions['D'].width = 15
    ws_skala.column_dimensions['G'].width = 15
    ws_skala.column_dimensions['J'].width = 20
    ws_skala.column_dimensions['K'].width = 25
    ws_skala.column_dimensions['M'].width = 20

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_file = os.path.join(script_dir, "Pamiętnik burz - poprawiony.xlsx")
    output_file = os.path.join(script_dir, "Pamiętnik burz - Nowe Skale.xlsx")
    
    wb = openpyxl.load_workbook(input_file)
    
    # Tworzenie arkusza konfiguracyjnego
    create_scale_sheet(wb)
    
    ws = wb["Burze"]

    # Kolumny docelowe i źródłowe
    for r in range(5, ws.max_row + 1):
        # VLOOKUP dla zagrożeń. Zwracamy uwagę na ISNUMBER żeby uniknąć błędu z "'-'"
        f_wiatr = f"IF(ISNUMBER(L{r}), VLOOKUP(L{r}, Skala!$A$2:$B$7, 2, TRUE), 0)"
        f_opad = f"IF(ISNUMBER(J{r}), VLOOKUP(J{r}, Skala!$D$2:$E$7, 2, TRUE), 0)"
        f_grad = f"IF(ISNUMBER(K{r}), VLOOKUP(K{r}, Skala!$G$2:$H$7, 2, TRUE), 0)"
        
        # Tornado punkty = Skala!$N$2
        f_tor = f'IF(OR(ISBLANK(M{r}), M{r}="-", M{r}="brak", M{r}=0), 0, Skala!$N$2)'
        
        # Suma punktów zagrożeń
        f_z_sum = f"({f_wiatr} + {f_opad} + {f_grad} + {f_tor})"
        
        # N - Ocena zagrożeń (z VLOOKUP po kategorii)
        formula_n = f'=VLOOKUP({f_z_sum}, Skala!$J$2:$K$7, 2, TRUE)'
        ws.cell(row=r, column=14, value=formula_n)
        
        # V - Ocena wyglądu (suma punktów z O do T -> następnie kategoria)
        f_w_sum = f"SUM(O{r}:T{r})"
        formula_v = f'=VLOOKUP({f_w_sum}, Skala!$J$2:$K$7, 2, TRUE)'
        ws.cell(row=r, column=22, value=formula_v)
        
        # X - Opłacalność: (punkty zagrożeń + punkty wyglądu) / KM * 100
        formula_x = f'=IF(ISNUMBER(W{r}), IF(W{r}>0, ({f_z_sum} + {f_w_sum}) / W{r} * 100, "-"), "-")'
        ws.cell(row=r, column=24, value=formula_x)

    wb.save(output_file)
    print("Workbook saved successfully as", output_file)

if __name__ == "__main__":
    main()
