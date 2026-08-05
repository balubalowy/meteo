# -*- coding: utf-8 -*-
"""
Skrypt uzupełniający Pamiętnik Burz danymi meteorologicznymi z Open-Meteo Archive API.
Wypełnia kolumny: G (T max), H (Td max), J (Deszcz), L (Wiatr).
"""

import openpyxl
import urllib.request
import json
import time
import os
from datetime import datetime

# ============================================================================
# SŁOWNIK LOKALIZACJI → WSPÓŁRZĘDNE (lat, lon)
# Współrzędne przybliżone do centrum miejscowości
# ============================================================================
LOCATION_COORDS = {
    # --- Opolszczyzna ---
    "Zagwiździe":       (50.583, 17.983),
    "Murów":            (50.667, 17.867),
    "Jaśkowice":        (50.600, 17.950),   # koło Zagwiździa
    "Prószków":         (50.567, 17.817),
    "Opole":            (50.675, 17.921),
    "Krapkowice":       (50.475, 17.963),
    "Ujazd":            (50.400, 17.783),   # koło Strzelec Opolskich
    "Grodków":          (50.700, 17.383),
    "Źlinice":          (50.683, 17.800),   # koło Opola
    "Łosiów":           (50.617, 17.533),   # koło Brzegu

    # --- Dolny Śląsk ---
    "Wrocław":          (51.108, 17.039),
    "Złotniki":         (51.167, 16.967),   # Złotniki koło Wrocławia
    "Prusice":          (51.317, 16.833),
    "Brzeg Dolny":      (51.267, 16.717),
    "Ząbkowice Śląskie": (50.583, 16.817),
    "Świdnica":         (50.850, 16.483),
    "Godzieszowa":      (51.033, 16.833),   # koło Wrocławia (SW)
    "Daniec":           (50.617, 17.150),   # koło Oławy

    # --- Śląsk ---
    "Gliwice":          (50.294, 18.671),
    "Jastrzębie-Zdrój": (49.950, 18.600),
    "Lubliniec":        (50.668, 18.683),
    "Poczesna":         (50.750, 19.100),   # koło Częstochowy
    "Poręba":           (50.500, 19.333),
    "Zawiercie":        (50.483, 19.417),
    "Myszków":          (50.567, 19.333),
    "Siewierz":         (50.467, 19.233),

    # --- Małopolska/Podhale ---
    "Biała":            (49.850, 19.000),   # Biała koło Bielska
    "Zator":            (49.983, 19.433),
    "Wilkowice":        (49.733, 19.083),   # koło Bielska-Białej
    "Jurgów":           (49.367, 20.117),   # Podhale

    # --- Centralna Polska ---
    "Warszawa":         (52.230, 21.012),
    "Kutno":            (52.230, 19.364),
    "Piotrków Trybunalski": (51.400, 19.700),
    "Ostrzeszów":       (51.417, 17.917),
    "Lututów":          (51.433, 18.517),

    # --- Zagranica ---
    "Paryż (FR)":      (48.857, 2.352),
}

# Aliasy (obsługa różnych kodowań/wariantów nazw z pliku Excel)
LOCATION_ALIASES = {
    # openpyxl może zwrócić nazwy z cp1250 artefaktami - dodajmy aliasy
}


def get_coords(location_name: str) -> tuple:
    """Zwraca (lat, lon) dla lokalizacji. Próbuje dopasować nazwę."""
    if not location_name:
        return None

    # Bezpośrednie dopasowanie
    if location_name in LOCATION_COORDS:
        return LOCATION_COORDS[location_name]

    # Aliasy
    if location_name in LOCATION_ALIASES:
        return LOCATION_COORDS[LOCATION_ALIASES[location_name]]

    # Próba fuzzy match - porównanie po usunięciu diakrytyków jest trudne,
    # więc spróbujmy porównać częściowo
    loc_lower = location_name.lower().strip()
    for key, coords in LOCATION_COORDS.items():
        if key.lower() in loc_lower or loc_lower in key.lower():
            return coords

    # Specjalne przypadki
    if "[16 burz]" in location_name:
        # Jurgów/Podhale - kontynuacja wyprawy z 5.08.2023
        return LOCATION_COORDS["Jurgów"]

    print(f"  ⚠️  Nie znaleziono współrzędnych dla: '{location_name}'")
    return None


def fetch_weather_data(lat: float, lon: float, date_str: str) -> dict:
    """
    Pobiera dane pogodowe z Open-Meteo Archive API.
    Zwraca dict z kluczami: t_max, td_max, precip, wind_gust
    """
    # API URL - daily + hourly dewpoint
    url = (
        f"https://archive-api.open-meteo.com/v1/archive"
        f"?latitude={lat}&longitude={lon}"
        f"&start_date={date_str}&end_date={date_str}"
        f"&daily=temperature_2m_max,precipitation_sum,wind_gusts_10m_max"
        f"&hourly=dewpoint_2m"
        f"&timezone=Europe/Warsaw"
    )

    try:
        req = urllib.request.Request(url, headers={"User-Agent": "PamietnikBurz/1.0"})
        with urllib.request.urlopen(req, timeout=30) as response:
            data = json.loads(response.read().decode("utf-8"))

        result = {
            "t_max": None,
            "td_max": None,
            "precip": None,
            "wind_gust": None,
        }

        # Daily data
        daily = data.get("daily", {})
        if daily.get("temperature_2m_max"):
            result["t_max"] = daily["temperature_2m_max"][0]
        if daily.get("precipitation_sum"):
            result["precip"] = daily["precipitation_sum"][0]
        if daily.get("wind_gusts_10m_max"):
            result["wind_gust"] = daily["wind_gusts_10m_max"][0]

        # Hourly dewpoint → compute daily max
        hourly = data.get("hourly", {})
        if hourly.get("dewpoint_2m"):
            dewpoints = [d for d in hourly["dewpoint_2m"] if d is not None]
            if dewpoints:
                result["td_max"] = round(max(dewpoints), 1)

        return result

    except Exception as e:
        print(f"  ❌ Błąd API dla {date_str} ({lat},{lon}): {e}")
        return None


def main():
    # Ścieżki plików
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_file = os.path.join(script_dir, "Pamiętnik burz.xlsx")
    output_file = os.path.join(script_dir, "Pamiętnik burz - uzupełniony.xlsx")

    print("=" * 60)
    print("  PAMIĘTNIK BURZ — Uzupełnianie danych meteorologicznych")
    print("  Tryb: DOBOWY (1 zapytanie per data, kopiowanie wartości)")
    print("=" * 60)
    print(f"\n📂 Plik wejściowy:  {input_file}")
    print(f"📂 Plik wyjściowy: {output_file}")

    # Wczytaj plik Excel
    wb = openpyxl.load_workbook(input_file)
    ws = wb["Burze"]

    print(f"\n📊 Arkusz 'Burze': {ws.max_row} wierszy, {ws.max_column} kolumn")

    # Zbierz wpisy: C=lp, D=data, E=godzina, F=lokalizacja
    # Wypełniamy: G=T max, H=Td max, J=Deszcz, L=Wiatr
    data_rows = []

    for row in range(5, ws.max_row + 1):
        lp = ws.cell(row=row, column=3).value  # C
        date_val = ws.cell(row=row, column=4).value  # D
        location = ws.cell(row=row, column=6).value  # F

        if date_val is None:
            continue

        # Konwertuj datę
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val)

        coords = get_coords(location)
        data_rows.append({
            "row": row,
            "lp": lp,
            "date": date_str,
            "location": location,
            "coords": coords,
        })

    print(f"\n🔍 Znaleziono {len(data_rows)} wpisów burz")

    # Grupuj po DACIE — dla każdej daty bierzemy współrzędne z PIERWSZEGO wpisu
    # Wszystkie burze z tego samego dnia dostaną identyczne dane pogodowe
    date_to_coords = {}  # date_str → (lat, lon)
    for dr in data_rows:
        if dr["date"] not in date_to_coords and dr["coords"]:
            date_to_coords[dr["date"]] = dr["coords"]

    unique_dates = sorted(date_to_coords.keys())
    print(f"📅 Unikalne daty: {len(unique_dates)}")
    print(f"\n{'='*60}")
    print("  Pobieranie danych z Open-Meteo Archive API...")
    print("  (1 zapytanie per datę → identyczne dane dla wszystkich burz danego dnia)")
    print(f"{'='*60}\n")

    # Pobierz dane — 1 zapytanie per datę
    cache = {}  # date_str → result
    for i, date_str in enumerate(unique_dates, 1):
        lat, lon = date_to_coords[date_str]
        print(f"  [{i}/{len(unique_dates)}] {date_str} ({lat:.3f}, {lon:.3f})", end="")

        result = fetch_weather_data(lat, lon, date_str)
        cache[date_str] = result

        if result:
            print(f"  → T={result['t_max']}°C, Td={result['td_max']}°C, "
                  f"P={result['precip']}mm, W={result['wind_gust']}km/h")
        else:
            print("  → BRAK DANYCH")

        # Rate limiting - 100ms between requests
        time.sleep(0.1)

    # Wypełnij komórki w arkuszu
    print(f"\n{'='*60}")
    print("  Zapisywanie danych do arkusza...")
    print(f"{'='*60}\n")

    filled = 0
    skipped = 0

    for dr in data_rows:
        row = dr["row"]
        date_str = dr["date"]

        result = cache.get(date_str)

        if not result:
            print(f"  ⏭️  Wiersz {row} ({dr['lp']}, {date_str}, {dr['location']}): "
                  f"POMINIĘTO (brak danych)")
            skipped += 1
            continue

        # G = T max (°C) — dobowy
        if result["t_max"] is not None:
            ws.cell(row=row, column=7, value=result["t_max"])

        # H = Td max (°C) — dobowy
        if result["td_max"] is not None:
            ws.cell(row=row, column=8, value=result["td_max"])

        # J = Deszcz (mm) — suma dobowa
        if result["precip"] is not None:
            ws.cell(row=row, column=10, value=result["precip"])

        # L = Wiatr (km/h) — max poryw dobowy
        if result["wind_gust"] is not None:
            ws.cell(row=row, column=12, value=result["wind_gust"])

        print(f"  ✅ Wiersz {row} ({dr['lp']}) {date_str} {dr['location']}: "
              f"T={result['t_max']}°C, Td={result['td_max']}°C, "
              f"P={result['precip']}mm, W={result['wind_gust']}km/h")
        filled += 1

    # Zapisz plik
    wb.save(output_file)

    print(f"\n{'='*60}")
    print(f"  GOTOWE!")
    print(f"{'='*60}")
    print(f"\n  ✅ Uzupełniono: {filled} wierszy")
    print(f"  ⏭️  Pominięto:   {skipped} wierszy")
    print(f"\n  📂 Zapisano do: {output_file}")


if __name__ == "__main__":
    main()

