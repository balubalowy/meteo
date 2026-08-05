import urllib.request
import json
import ssl
import re
import os
import time
from datetime import datetime
import openpyxl

def get_max_threat_level(date_str: str) -> str:
    """
    Pobiera najwyższy stopień zagrożenia dla danej daty.
    date_str: YYYY-MM-DD
    """
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    
    # 06.06.2024 -> 6.06.2024 (sometimes they use single digit)
    day = date_obj.day
    month = date_obj.month
    year = date_obj.year
    date_query1 = f"{day:02d}.{month:02d}.{year}"
    date_query2 = f"{day}.{month:02d}.{year}"
    
    search_query = urllib.parse.quote(date_query2)
    url = f"https://lowcyburz.pl/wp-json/wp/v2/posts?search={search_query}"
    
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE

    max_threat = -1

    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, context=ctx, timeout=15) as response:
            data = json.loads(response.read().decode('utf-8'))
            
            for post in data:
                title = post['title']['rendered'].lower()
                # Check if title is actually a forecast for this date
                if 'prognoza' in title and (date_query1 in title or date_query2 in title):
                    content = post['content']['rendered']
                    
                    # Znajdź wszystkie stopnie: np. "2. STOPIEŃ", "3 STOPIEŃ"
                    matches = re.findall(r'([123])\.?\s*STOPIE[NŃ]', content, re.IGNORECASE)
                    
                    if matches:
                        threats = [int(m) for m in matches]
                        current_max = max(threats)
                        if current_max > max_threat:
                            max_threat = current_max
                    else:
                        # Jeśli nie ma stopni, ale jest prognoza burz (Możliwość burz)
                        if max_threat == -1:
                            max_threat = 0
            
    except Exception as e:
        print(f"Error fetching {date_str}: {e}")
        return "?"

    if max_threat == 3: return "3. stopień"
    if max_threat == 2: return "2. stopień"
    if max_threat == 1: return "1. stopień"
    if max_threat == 0: return "Możliwość burz"
    
    return "Brak/Brak danych"

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_file = os.path.join(script_dir, "Pamiętnik burz - uzupełniony.xlsx")
    
    wb = openpyxl.load_workbook(input_file)
    ws = wb["Burze"]
    
    # 1. Zbierz unikalne daty
    dates = set()
    for row in range(5, ws.max_row + 1):
        date_val = ws.cell(row=row, column=4).value
        if date_val:
            if isinstance(date_val, datetime):
                dates.add(date_val.strftime("%Y-%m-%d"))
            else:
                dates.add(str(date_val)[:10])

    # 2. Pobierz prognozy
    cache = {}
    print(f"Pobieranie prognoz PŁB dla {len(dates)} dat...")
    for date_str in sorted(dates):
        print(f"Pobieranie dla: {date_str}...", end=" ")
        threat = get_max_threat_level(date_str)
        cache[date_str] = threat
        print(f"-> {threat}")
        time.sleep(0.5)

    # 3. Zapisz do excela
    filled = 0
    for row in range(5, ws.max_row + 1):
        date_val = ws.cell(row=row, column=4).value
        if not date_val: continue
        
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val)[:10]
            
        threat = cache.get(date_str, "")
        if threat and threat != "?":
            ws.cell(row=row, column=9, value=threat) # Kolumna I
            filled += 1
            
    wb.save(input_file)
    print(f"Zaktualizowano wierszy: {filled}")

if __name__ == '__main__':
    main()
