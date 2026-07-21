import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

def create_excel_calculator():
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # ARKUSZ 1: Przelicznik Jednostek Meteorologicznych
    # ----------------------------------------------------
    ws_units = wb.active
    ws_units.title = "Przelicznik Jednostek"
    ws_units.views.sheetView[0].showGridLines = True

    # Nagłówek
    ws_units.merge_cells("A1:E1")
    ws_units["A1"] = "PRZELICZNIK JEDNOSTEK METEOROLOGICZNYCH"
    ws_units["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws_units["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws_units["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Tabela 1: Prędkość wiatru
    ws_units["A3"] = "PRĘDKOŚĆ WIATRU"
    ws_units["A3"].font = Font(name="Calibri", size=12, bold=True, color="1F4E79")
    
    headers_wind = ["Wartość wejściowa", "Jednostka wej.", "Wynik km/h", "Wynik m/s", "Wynik węzły (kt)", "Źródło przelicznika"]
    for col_idx, text in enumerate(headers_wind, 1):
        cell = ws_units.cell(row=4, column=col_idx, value=text)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5597")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Przykładowe wiersze przelicznika wiatru
    # Wzory: 
    # 1 kt = 1.852 km/h (allmetsat)
    # 1 m/s = 3.6 km/h (allmetsat)
    wind_data = [
        (100, "km/h", "=A5", "=A5/3.6", "=A5/1.852", "1 m/s = 3.6 km/h; 1 kt = 1.852 km/h (allmetsat)"),
        (25, "m/s", "=A6*3.6", "=A6", "=(A6*3.6)/1.852", "1 m/s = 3.6 km/h; 1 kt = 1.852 km/h (allmetsat)"),
        (50, "kt", "=A7*1.852", "=(A7*1.852)/3.6", "=A7", "1 kt = 1.852 km/h (allmetsat)"),
    ]

    for row_idx, row_val in enumerate(wind_data, 5):
        ws_units.cell(row=row_idx, column=1, value=row_val[0]).alignment = Alignment(horizontal="center")
        ws_units.cell(row=row_idx, column=2, value=row_val[1]).alignment = Alignment(horizontal="center")
        ws_units.cell(row=row_idx, column=3, value=row_val[2]).alignment = Alignment(horizontal="center")
        ws_units.cell(row=row_idx, column=4, value=row_val[3]).alignment = Alignment(horizontal="center")
        ws_units.cell(row=row_idx, column=5, value=row_val[4]).alignment = Alignment(horizontal="center")
        src_cell = ws_units.cell(row=row_idx, column=6, value=row_val[5])
        src_cell.font = Font(italic=True, size=9, color="595959")
        src_cell.comment = Comment("Współczynniki konwersji ze źródła allmetsat:\n- 1 kt = 1.852 km/h\n- 1 m/s = 3.6 km/h", "Antigravity Meteo Agent")

    # Tabela 2: Ciśnienie atmosferyczne
    ws_units["A10"] = "CIŚNIENIE ATMOSFERYCZNE"
    ws_units["A10"].font = Font(name="Calibri", size=12, bold=True, color="1F4E79")

    headers_press = ["Wartość wejściowa", "Jednostka wej.", "Wynik hPa", "Wynik inHg", "Wynik atm", "Źródło przelicznika"]
    for col_idx, text in enumerate(headers_press, 1):
        cell = ws_units.cell(row=11, column=col_idx, value=text)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5597")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    press_data = [
        (1013.25, "hPa", "=A12", "=A12/33.86389", "=A12/1013.25", "1 inHg = 33.86 hPa; 1 atm = 1013.25 hPa (allmetsat)"),
        (29.92, "inHg", "=A13*33.86389", "=A13", "=(A13*33.86389)/1013.25", "1 inHg = 33.86 hPa (allmetsat)"),
        (1.0, "atm", "=A14*1013.25", "=(A14*1013.25)/33.86389", "=A14", "1 atm = 1013.25 hPa (allmetsat)"),
    ]

    for row_idx, row_val in enumerate(press_data, 12):
        ws_units.cell(row=row_idx, column=1, value=row_val[0]).alignment = Alignment(horizontal="center")
        ws_units.cell(row=row_idx, column=2, value=row_val[1]).alignment = Alignment(horizontal="center")
        ws_units.cell(row=row_idx, column=3, value=row_val[2]).alignment = Alignment(horizontal="center")
        ws_units.cell(row=row_idx, column=4, value=row_val[3]).alignment = Alignment(horizontal="center")
        ws_units.cell(row=row_idx, column=5, value=row_val[4]).alignment = Alignment(horizontal="center")
        src_cell = ws_units.cell(row=row_idx, column=6, value=row_val[5])
        src_cell.font = Font(italic=True, size=9, color="595959")
        src_cell.comment = Comment("Współczynniki konwersji ze źródła allmetsat:\n- 1 inHg = 33.86 hPa\n- 1 atm = 1013.25 hPa", "Antigravity Meteo Agent")


    # ----------------------------------------------------
    # ARKUSZ 2: Skala IF (International Fujita)
    # ----------------------------------------------------
    ws_if = wb.create_sheet(title="Skala IF (Fujita)")
    ws_if.views.sheetView[0].showGridLines = True

    ws_if.merge_cells("A1:G1")
    ws_if["A1"] = "KLASYFIKACJA SZKÓD W SKALI INTERNATIONAL FUJITA (IF-SCALE)"
    ws_if["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws_if["A1"].fill = PatternFill("solid", fgColor="C00000")
    ws_if["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers_if = ["Stopień IF", "Wiatr min (km/h)", "Wiatr max (km/h)", "Wiatr min (m/s)", "Wiatr max (m/s)", "Wiatr min (kt)", "Opis Zniszczeń i Kryteria (ESSL)"]
    for col_idx, text in enumerate(headers_if, 1):
        cell = ws_if.cell(row=3, column=col_idx, value=text)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C00000")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if_table = [
        ("IF0", 90, 125, "=B4/3.6", "=C4/3.6", "=B4/1.852", "Słabe zniszczenia: łamanie gałęzi, uszkodzenia lekkich dachówek, ogrodzeń."),
        ("IF0.5", 125, 155, "=B5/3.6", "=C5/3.6", "=B5/1.852", "Małe uszkodzenia: uszkodzenia poszyć dachowych, łamanie konarów, pojedyncze drzewa."),
        ("IF1", 155, 190, "=B6/3.6", "=C6/3.6", "=B6/1.852", "Umiarkowane zniszczenia: zrywanie dachów z lekkich budynków, przewracanie przyczep."),
        ("IF1.5", 190, 235, "=B7/3.6", "=C7/3.6", "=B7/1.852", "Znaczne uszkodzenia: zrywanie dachów domów mieszkalnych, masowo powalone drzewa."),
        ("IF2", 235, 280, "=B8/3.6", "=C8/3.6", "=B8/1.852", "Ciężkie zniszczenia: zawalenie ścian osłonowych, odkształcenia konstrukcji stalowych."),
        ("IF2.5", 280, 330, "=B9/3.6", "=C9/3.6", "=B9/1.852", "Bardzo ciężkie zniszczenia: zniszczenie wyższych pięter budynków, odkorowywanie drzew."),
        ("IF3", 330, 385, "=B10/3.6", "=C10/3.6", "=B10/1.852", "Niszczycielskie skutki: zrównanie z ziemią budynków murowanych, unoszenie pojazdów."),
        ("IF4", 385, 450, "=B11/3.6", "=C11/3.6", "=B11/1.852", "Katastrofalne zniszczenia: budynki zrównane z ziemią, ciśnięte ciężkie obiekty."),
        ("IF5", 450, 600, "=B12/3.6", "=C12/3.6", "=B12/1.852", "Ekstremalna destrukcja: starcie z fundamentów, odklejanie asfaltu z dróg."),
    ]

    for row_idx, row_val in enumerate(if_table, 4):
        ws_if.cell(row=row_idx, column=1, value=row_val[0]).alignment = Alignment(horizontal="center")
        ws_if.cell(row=row_idx, column=2, value=row_val[1]).alignment = Alignment(horizontal="center")
        ws_if.cell(row=row_idx, column=3, value=row_val[2]).alignment = Alignment(horizontal="center")
        ws_if.cell(row=row_idx, column=4, value=row_val[3]).alignment = Alignment(horizontal="center")
        ws_if.cell(row=row_idx, column=5, value=row_val[4]).alignment = Alignment(horizontal="center")
        ws_if.cell(row=row_idx, column=6, value=row_val[5]).alignment = Alignment(horizontal="center")
        desc_cell = ws_if.cell(row=row_idx, column=7, value=row_val[6])
        desc_cell.alignment = Alignment(horizontal="left")
        desc_cell.comment = Comment("Źródło: ESSL (European Severe Storms Laboratory) - International Fujita Scale Draft Standard. Przeliczniki: 1 kt = 1.852 km/h; 1 m/s = 3.6 km/h (allmetsat).", "Antigravity Meteo Agent")

    # ----------------------------------------------------
    # ARKUSZ 3: Kalkulator LCL & Odczuwalnej
    # ----------------------------------------------------
    ws_calc = wb.create_sheet(title="Kalkulator LCL i Odczuwalnej")
    ws_calc.views.sheetView[0].showGridLines = True

    ws_calc.merge_cells("A1:E1")
    ws_calc["A1"] = "KALKULATOR METEOROLOGICZNY (LCL & TEMPERATURA ODCZUWANA)"
    ws_calc["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws_calc["A1"].fill = PatternFill("solid", fgColor="385723")
    ws_calc["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Sekcja LCL
    ws_calc["A3"] = "1. Estymacja wysokości podstawy chmur (LCL - Espy formula)"
    ws_calc["A3"].font = Font(bold=True, size=11, color="385723")

    ws_calc["A4"] = "Temperatura (T) [°C]:"
    ws_calc["B4"] = 25.0
    ws_calc["A5"] = "Punkt rosy (Td) [°C]:"
    ws_calc["B5"] = 18.0
    ws_calc["A6"] = "Szacowana wysokość LCL [m n.p.g.]:"
    ws_calc["B6"] = "=125*(B4-B5)"
    ws_calc["B6"].font = Font(bold=True, color="385723")
    
    ws_calc["C6"] = "Wzór Espy'ego: h_LCL = 125 * (T - Td) [m]. Źródło: NOAA / Espy."
    ws_calc["C6"].font = Font(italic=True, size=9, color="595959")
    ws_calc["C6"].comment = Comment("Zależność Espy'ego wyznacza przybliżony poziom kondensacji z podnoszenia (LCL) dla cząstki konwekcyjnej przy ziemi.", "Antigravity Meteo Agent")

    # Auto-adjust column widths
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if not val_str.startswith("="):
                    max_len = max(max_len, len(val_str))
                else:
                    max_len = max(max_len, 10)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    os.makedirs("e:/meteo/kalkulatory_excel", exist_ok=True)
    out_path = "e:/meteo/kalkulatory_excel/Kalkulator_Meteo_Skala_IF.xlsx"
    wb.save(out_path)
    print(f"Pomyślnie utworzono plik Excel: {out_path}")

if __name__ == "__main__":
    create_excel_calculator()
